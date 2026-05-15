#!/usr/bin/env python3
"""Aggregate per-cluster sub-agent results into a final Excel evaluation report.

Usage:
    python aggregate_scores.py \
        --rubric rubric.json \
        --clusters-dir outputs/ \
        --output evaluation_report.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_rubric(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_cluster_results(clusters_dir: Path, rubric: dict) -> dict[str, dict]:
    """Load each cluster_<id>.json into a dict keyed by cluster_id."""
    results: dict[str, dict] = {}
    expected = {c["id"] for c in rubric["clusters"]}
    for path in clusters_dir.glob("cluster_*.json"):
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        results[data["cluster_id"]] = data
    missing = expected - results.keys()
    if missing:
        print(f"WARNING: Missing cluster results for: {sorted(missing)}", file=sys.stderr)
    return results


def compute_final_grade(rubric: dict, results: dict[str, dict]) -> tuple[float, float]:
    total = sum(r.get("cluster_score", 0) for r in results.values())
    max_total = rubric.get("total_points", sum(c["weight"] for c in rubric["clusters"]))
    return total, max_total


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def build_workbook(rubric: dict, results: dict[str, dict], total: float, max_total: float) -> Workbook:
    wb = Workbook()

    # --- Sheet 1: Samenvatting ---
    s1 = wb.active
    s1.title = "Samenvatting"
    s1["A1"] = "Evaluatierapport"
    s1["A1"].font = Font(bold=True, size=14)
    s1["A2"] = f"Opdracht: {rubric.get('title', '-')}"

    grade_10 = round(total / max_total * 10, 1) if max_total else 0
    s1["A4"] = "Eindcijfer (1-10)"
    s1["B4"] = grade_10
    s1["A4"].font = Font(bold=True)
    s1["B4"].font = Font(bold=True, size=14, color="1F4E78")

    s1["A5"] = "Totaal punten"
    s1["B5"] = f"{round(total, 1)} / {round(max_total, 1)}"

    # Cluster summary table
    s1["A7"] = "Cluster"
    s1["B7"] = "Score"
    s1["C7"] = "Max"
    s1["D7"] = "%"
    for col in ["A7", "B7", "C7", "D7"]:
        style_header(s1[col])

    row = 8
    for cluster in rubric["clusters"]:
        res = results.get(cluster["id"], {})
        score = res.get("cluster_score", 0)
        cmax = res.get("cluster_max", cluster["weight"])
        pct = round(score / cmax * 100, 0) if cmax else 0
        s1.cell(row=row, column=1, value=cluster["name"]).border = BORDER
        s1.cell(row=row, column=2, value=round(score, 1)).border = BORDER
        s1.cell(row=row, column=3, value=round(cmax, 1)).border = BORDER
        s1.cell(row=row, column=4, value=f"{pct}%").border = BORDER
        row += 1

    for col_letter, width in [("A", 28), ("B", 10), ("C", 10), ("D", 10)]:
        s1.column_dimensions[col_letter].width = width

    # --- Sheet 2: Detail per criterium ---
    s2 = wb.create_sheet("Detail per criterium")
    headers = ["Cluster", "Criterium", "Score", "Max", "Niveau", "Motivering"]
    for idx, h in enumerate(headers, start=1):
        c = s2.cell(row=1, column=idx, value=h)
        style_header(c)

    row = 2
    for cluster in rubric["clusters"]:
        res = results.get(cluster["id"], {})
        for crit in res.get("criteria", []):
            s2.cell(row=row, column=1, value=cluster["name"]).border = BORDER
            # Look up the criterion's description from the rubric
            desc = next(
                (c["description"] for c in cluster["criteria"] if c["id"] == crit["id"]),
                crit["id"],
            )
            s2.cell(row=row, column=2, value=desc).border = BORDER
            s2.cell(row=row, column=3, value=round(crit.get("achieved_score", 0), 1)).border = BORDER
            s2.cell(row=row, column=4, value=round(crit.get("max_score", 0), 1)).border = BORDER
            s2.cell(row=row, column=5, value=crit.get("level_label", "-")).border = BORDER
            motiv_cell = s2.cell(row=row, column=6, value=crit.get("motivation", ""))
            motiv_cell.alignment = Alignment(wrap_text=True, vertical="top")
            motiv_cell.border = BORDER
            row += 1

    for col_letter, width in [("A", 20), ("B", 42), ("C", 8), ("D", 8), ("E", 14), ("F", 70)]:
        s2.column_dimensions[col_letter].width = width

    # --- Sheet 3: Rubric (reproduceerbaarheid) ---
    s3 = wb.create_sheet("Rubric")
    s3["A1"] = "Gebruikte rubric"
    s3["A1"].font = Font(bold=True, size=12)
    s3["A3"] = json.dumps(rubric, indent=2, ensure_ascii=False)
    s3["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    s3.column_dimensions["A"].width = 120

    return wb


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--rubric", required=True, type=Path)
    parser.add_argument("--clusters-dir", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    rubric = load_rubric(args.rubric)
    results = load_cluster_results(args.clusters_dir, rubric)
    total, max_total = compute_final_grade(rubric, results)
    wb = build_workbook(rubric, results, total, max_total)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    grade_10 = round(total / max_total * 10, 1) if max_total else 0
    print(f"Eindcijfer: {grade_10} ({round(total,1)}/{round(max_total,1)} punten)")
    print(f"Rapport opgeslagen: {args.output}")


if __name__ == "__main__":
    main()
